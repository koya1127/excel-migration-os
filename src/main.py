from __future__ import annotations

import argparse
import csv
import json
import os
import re
import zipfile
from dataclasses import asdict, dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable, List, Optional

from dotenv import load_dotenv
from openpyxl import load_workbook

load_dotenv()

EXCEL_EXTENSIONS = {".xls", ".xlsx", ".xlsm"}
VOLATILE_FUNCTIONS = {
    "NOW(",
    "TODAY(",
    "RAND(",
    "RANDBETWEEN(",
    "OFFSET(",
    "INDIRECT(",
    "CELL(",
    "INFO(",
}
# Excel functions that do not exist or behave differently in Google Sheets.
# Categories: "missing" = no equivalent, "partial" = exists but limited,
#             "check" = usually works but needs manual verification.
INCOMPATIBLE_FUNCTIONS: dict[str, str] = {
    # --- missing: no Google Sheets equivalent ---
    "AGGREGATE(": "missing",
    "CUBEVALUE(": "missing",
    "CUBEMEMBER(": "missing",
    "CUBESET(": "missing",
    "CUBERANKEDMEMBER(": "missing",
    "CUBEKPIMEMBER(": "missing",
    "CUBESETCOUNT(": "missing",
    "CUBEMEMBERPROPERTY(": "missing",
    "CALL(": "missing",
    "REGISTER.ID(": "missing",
    "RTD(": "missing",
    "SQL.REQUEST(": "missing",
    "EUROCONVERT(": "missing",
    "WEBSERVICE(": "missing",
    "FILTERXML(": "missing",
    "PHONETIC(": "missing",
    "JIS(": "missing",
    "ASC(": "missing",
    "BAHTTEXT(": "missing",
    # --- partial: exists but limited or different behaviour ---
    "XLOOKUP(": "partial",
    "XMATCH(": "partial",
    "LET(": "partial",
    "LAMBDA(": "partial",
    "GETPIVOTDATA(": "partial",
    "INFO(": "partial",
    "CELL(": "partial",
    "ERROR.TYPE(": "partial",
    # --- check: usually fine but worth flagging ---
    "SUBTOTAL(": "check",
    "INDIRECT(": "check",
    "OFFSET(": "check",
}

DRIVE_SCOPES = ["https://www.googleapis.com/auth/drive.file"]
SCRIPT_SCOPES = [
    "https://www.googleapis.com/auth/drive.file",
    "https://www.googleapis.com/auth/script.projects",
]


@dataclass
class FileReport:
    path: str
    extension: str
    size_bytes: int
    modified_utc: str
    has_macro: bool
    vba_module_count: int
    sheet_count: int
    formula_count: int
    volatile_formula_count: int
    named_range_count: int
    external_link_count: int
    incompatible_function_count: int
    risk_score: int
    notes: str


@dataclass
class GroupSummary:
    group_name: str
    file_count: int
    total_size_bytes: int
    macro_file_count: int
    total_vba_modules: int
    avg_risk_score: float
    max_risk_score: int
    total_formulas: int
    total_incompatible_functions: int
    migration_difficulty: str
    file_indices: List[int]


@dataclass
class ScanReport:
    generated_utc: str
    input_root: str
    file_count: int
    files: List[FileReport]
    group_by: str = "none"
    groups: List[GroupSummary] = field(default_factory=list)


@dataclass
class VbaModule:
    source_file: str
    module_name: str
    module_type: str
    code_lines: int
    output_path: str


@dataclass
class FormControl:
    source_file: str
    sheet_name: str
    control_name: str
    control_type: str
    label: str
    macro: str
    anchor: str


@dataclass
class ExtractReport:
    generated_utc: str
    input_root: str
    file_count: int
    module_count: int
    modules: List[VbaModule]
    controls: List[FormControl]


@dataclass
class ConvertResult:
    source_bas: str
    module_name: str
    module_type: str
    output_gs: str
    status: str
    error: str


@dataclass
class ConvertReport:
    generated_utc: str
    input_root: str
    total: int
    success: int
    failed: int
    results: List[ConvertResult]


@dataclass
class UploadResult:
    local_path: str
    file_name: str
    source_mime_type: str
    target_mime_type: str
    drive_file_id: str
    drive_web_view_link: str
    status: str
    error: str


@dataclass
class DeployReport:
    generated_utc: str
    spreadsheet_id: str
    script_id: str
    file_count: int
    files_deployed: List[str]


@dataclass
class UploadReport:
    generated_utc: str
    input_root: str
    output_root: str
    file_count: int
    success_count: int
    failure_count: int
    converted_to_sheets: bool
    drive_folder_id: str
    files: List[UploadResult]


def find_excel_files(root: Path) -> Iterable[Path]:
    if root.is_file():
        if root.suffix.lower() in EXCEL_EXTENSIONS:
            yield root
        return
    for file_path in root.rglob("*"):
        if file_path.is_file() and file_path.suffix.lower() in EXCEL_EXTENSIONS:
            yield file_path


def contains_vba_project(file_path: Path) -> bool:
    if file_path.suffix.lower() == ".xlsm":
        return True
    if file_path.suffix.lower() != ".xlsx":
        return False

    try:
        with zipfile.ZipFile(file_path, "r") as zf:
            names = set(zf.namelist())
        return "xl/vbaProject.bin" in names
    except (zipfile.BadZipFile, OSError):
        return False


def try_count_vba_modules(file_path: Path) -> int:
    try:
        from oletools.olevba import VBA_Parser  # type: ignore
    except Exception:
        return -1

    try:
        parser = VBA_Parser(str(file_path))
        if not parser.detect_vba_macros():
            return 0
        module_count = 0
        for _ in parser.extract_macros():
            module_count += 1
        parser.close()
        return module_count
    except Exception:
        return -1


def count_external_links(workbook) -> int:
    links = getattr(workbook, "_external_links", None)
    if links is None:
        return 0
    return len(links)


def analyze_openxml_file(file_path: Path, has_macro: bool, vba_module_count: int) -> FileReport:
    workbook = load_workbook(filename=file_path, data_only=False, keep_vba=True)
    try:
        formula_count = 0
        volatile_formula_count = 0
        incompatible_funcs_found: set[str] = set()
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(values_only=False):
                for cell in row:
                    value = cell.value
                    if isinstance(value, str) and value.startswith("="):
                        formula_count += 1
                        upper_formula = value.upper()
                        if any(fn in upper_formula for fn in VOLATILE_FUNCTIONS):
                            volatile_formula_count += 1
                        for fn, cat in INCOMPATIBLE_FUNCTIONS.items():
                            if fn in upper_formula:
                                incompatible_funcs_found.add(fn.rstrip("("))

        named_range_count = len(workbook.defined_names)
        external_link_count = count_external_links(workbook)
        sheet_count = len(workbook.sheetnames)
    finally:
        workbook.close()

    risk_score = 0
    notes = []
    if has_macro:
        risk_score += 40
        notes.append("マクロ有り")
    if volatile_formula_count > 0:
        risk_score += 15
        notes.append("揮発性の数式あり")
    if external_link_count > 0:
        risk_score += 20
        notes.append("外部リンクあり")
    if formula_count > 10000:
        risk_score += 10
        notes.append("数式が多い")
    if incompatible_funcs_found:
        risk_score += 25
        notes.append(f"非互換関数: {', '.join(sorted(incompatible_funcs_found))}")

    stat = file_path.stat()
    modified = datetime.fromtimestamp(stat.st_mtime, tz=timezone.utc).isoformat()

    return FileReport(
        path=str(file_path),
        extension=file_path.suffix.lower(),
        size_bytes=stat.st_size,
        modified_utc=modified,
        has_macro=has_macro,
        vba_module_count=vba_module_count,
        sheet_count=sheet_count,
        formula_count=formula_count,
        volatile_formula_count=volatile_formula_count,
        named_range_count=named_range_count,
        external_link_count=external_link_count,
        incompatible_function_count=len(incompatible_funcs_found),
        risk_score=min(risk_score, 100),
        notes="; ".join(notes) if notes else "リスク低",
    )


def analyze_legacy_xls(file_path: Path, has_macro: bool, vba_module_count: int) -> FileReport:
    stat = file_path.stat()
    modified = datetime.fromtimestamp(stat.st_mtime, tz=timezone.utc).isoformat()

    risk_score = 30
    notes = "旧形式(.xls); .xlsxに変換すると詳細分析可能"
    if has_macro:
        risk_score += 40
        notes += "; マクロ有りの可能性"

    return FileReport(
        path=str(file_path),
        extension=file_path.suffix.lower(),
        size_bytes=stat.st_size,
        modified_utc=modified,
        has_macro=has_macro,
        vba_module_count=vba_module_count,
        sheet_count=0,
        formula_count=0,
        volatile_formula_count=0,
        named_range_count=0,
        external_link_count=0,
        incompatible_function_count=0,
        risk_score=min(risk_score, 100),
        notes=notes,
    )


def analyze_file(file_path: Path) -> FileReport:
    has_macro = contains_vba_project(file_path)
    vba_module_count = try_count_vba_modules(file_path)

    if file_path.suffix.lower() == ".xls":
        return analyze_legacy_xls(file_path, has_macro, vba_module_count)

    try:
        return analyze_openxml_file(file_path, has_macro, vba_module_count)
    except Exception as exc:
        stat = file_path.stat()
        modified = datetime.fromtimestamp(stat.st_mtime, tz=timezone.utc).isoformat()
        return FileReport(
            path=str(file_path),
            extension=file_path.suffix.lower(),
            size_bytes=stat.st_size,
            modified_utc=modified,
            has_macro=has_macro,
            vba_module_count=vba_module_count,
            sheet_count=0,
            formula_count=0,
            volatile_formula_count=0,
            named_range_count=0,
            external_link_count=0,
            incompatible_function_count=0,
            risk_score=80,
            notes=f"Analysis failed: {exc}",
        )


def write_json_report(report: ScanReport, output_dir: Path) -> Path:
    output_path = output_dir / "report.json"
    payload = {
        "generated_utc": report.generated_utc,
        "input_root": report.input_root,
        "file_count": report.file_count,
        "group_by": report.group_by,
        "groups": [asdict(g) for g in report.groups],
        "files": [asdict(item) for item in report.files],
    }
    output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return output_path


def write_csv_report(report: ScanReport, output_dir: Path) -> Path:
    output_path = output_dir / "report.csv"
    with output_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=list(asdict(report.files[0]).keys())
            if report.files
            else list(
                asdict(
                    FileReport(
                        path="",
                        extension="",
                        size_bytes=0,
                        modified_utc="",
                        has_macro=False,
                        vba_module_count=0,
                        sheet_count=0,
                        formula_count=0,
                        volatile_formula_count=0,
                        named_range_count=0,
                        external_link_count=0,
                        incompatible_function_count=0,
                        risk_score=0,
                        notes="",
                    )
                ).keys()
            ),
        )
        writer.writeheader()
        for item in report.files:
            writer.writerow(asdict(item))
    return output_path


def write_group_csv_report(report: ScanReport, output_dir: Path) -> Path:
    output_path = output_dir / "group_summary.csv"
    fieldnames = [
        "group_name", "file_count", "total_size_bytes", "macro_file_count",
        "total_vba_modules", "avg_risk_score", "max_risk_score",
        "total_formulas", "total_incompatible_functions", "migration_difficulty",
    ]
    with output_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for g in report.groups:
            row = asdict(g)
            del row["file_indices"]
            writer.writerow(row)
    return output_path


def _extract_group_name_prefix(file_path: str) -> str:
    """Extract group name from filename prefix before first '_'."""
    name = Path(file_path).stem
    if "_" in name:
        return name.split("_", 1)[0]
    return "(未分類)"


def _extract_group_name_subfolder(file_path: str, input_root: str) -> str:
    """Extract group name from immediate subfolder under input_root."""
    try:
        rel = Path(file_path).relative_to(input_root)
    except ValueError:
        return "(未分類)"
    parts = rel.parts
    if len(parts) <= 1:
        return str(Path(input_root).name)
    return parts[0]


def _classify_migration_difficulty(
    max_risk: int, has_macro: bool, has_incompatible: bool
) -> str:
    if max_risk >= 70 or has_incompatible:
        return "Hard"
    if max_risk >= 40 or has_macro:
        return "Medium"
    return "Easy"


def build_group_summaries(
    files: List[FileReport], group_by: str, input_root: str
) -> List[GroupSummary]:
    if group_by == "none" or not files:
        return []

    groups: dict[str, list[int]] = {}
    for i, f in enumerate(files):
        if group_by == "prefix":
            name = _extract_group_name_prefix(f.path)
        else:
            name = _extract_group_name_subfolder(f.path, input_root)
        groups.setdefault(name, []).append(i)

    summaries: List[GroupSummary] = []
    for group_name, indices in sorted(groups.items()):
        group_files = [files[i] for i in indices]
        total_size = sum(f.size_bytes for f in group_files)
        macro_count = sum(1 for f in group_files if f.has_macro)
        vba_modules = sum(
            max(f.vba_module_count, 0) for f in group_files
        )
        risks = [f.risk_score for f in group_files]
        avg_risk = round(sum(risks) / len(risks), 1) if risks else 0.0
        max_risk = max(risks) if risks else 0
        total_formulas = sum(f.formula_count for f in group_files)
        total_incompat = sum(
            f.incompatible_function_count for f in group_files
        )
        difficulty = _classify_migration_difficulty(
            max_risk, macro_count > 0, total_incompat > 0
        )
        summaries.append(
            GroupSummary(
                group_name=group_name,
                file_count=len(group_files),
                total_size_bytes=total_size,
                macro_file_count=macro_count,
                total_vba_modules=vba_modules,
                avg_risk_score=avg_risk,
                max_risk_score=max_risk,
                total_formulas=total_formulas,
                total_incompatible_functions=total_incompat,
                migration_difficulty=difficulty,
                file_indices=indices,
            )
        )
    return summaries


def run_scan(input_root: Path, output_dir: Path, group_by: str = "none") -> ScanReport:
    files = list(find_excel_files(input_root))
    reports = [analyze_file(path) for path in files]

    groups = build_group_summaries(reports, group_by, str(input_root))

    scan_report = ScanReport(
        generated_utc=datetime.now(tz=timezone.utc).isoformat(),
        input_root=str(input_root),
        file_count=len(reports),
        files=reports,
        group_by=group_by,
        groups=groups,
    )

    output_dir.mkdir(parents=True, exist_ok=True)
    write_json_report(scan_report, output_dir)
    write_csv_report(scan_report, output_dir)
    if groups:
        write_group_csv_report(scan_report, output_dir)
    return scan_report


def load_drive_service(credentials_path: Path, token_path: Path):
    from google.auth.transport.requests import Request
    from google.oauth2.credentials import Credentials
    from google_auth_oauthlib.flow import InstalledAppFlow
    from googleapiclient.discovery import build

    creds = None
    if token_path.exists():
        creds = Credentials.from_authorized_user_file(str(token_path), DRIVE_SCOPES)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(str(credentials_path), DRIVE_SCOPES)
            creds = flow.run_local_server(port=0)
        token_path.parent.mkdir(parents=True, exist_ok=True)
        token_path.write_text(creds.to_json(), encoding="utf-8")

    return build("drive", "v3", credentials=creds, cache_discovery=False)


def get_upload_mime_type(file_path: Path) -> str:
    ext = file_path.suffix.lower()
    if ext == ".xls":
        return "application/vnd.ms-excel"
    if ext == ".xlsx":
        return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    if ext == ".xlsm":
        return "application/vnd.ms-excel.sheet.macroEnabled.12"
    return "application/octet-stream"


def write_upload_json(report: UploadReport, output_dir: Path) -> Path:
    output_path = output_dir / "upload_report.json"
    output_path.write_text(
        json.dumps(
            {
                "generated_utc": report.generated_utc,
                "input_root": report.input_root,
                "output_root": report.output_root,
                "file_count": report.file_count,
                "success_count": report.success_count,
                "failure_count": report.failure_count,
                "converted_to_sheets": report.converted_to_sheets,
                "drive_folder_id": report.drive_folder_id,
                "files": [asdict(f) for f in report.files],
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    return output_path


def write_upload_csv(report: UploadReport, output_dir: Path) -> Path:
    output_path = output_dir / "upload_report.csv"
    with output_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=list(asdict(report.files[0]).keys())
            if report.files
            else list(
                asdict(
                    UploadResult(
                        local_path="",
                        file_name="",
                        source_mime_type="",
                        target_mime_type="",
                        drive_file_id="",
                        drive_web_view_link="",
                        status="",
                        error="",
                    )
                ).keys()
            ),
        )
        writer.writeheader()
        for item in report.files:
            writer.writerow(asdict(item))
    return output_path


def run_upload(
    input_root: Path,
    output_dir: Path,
    credentials_path: Path,
    token_path: Path,
    drive_folder_id: Optional[str],
    convert_to_sheets: bool,
) -> UploadReport:
    from googleapiclient.http import MediaFileUpload

    service = load_drive_service(credentials_path, token_path)
    files = list(find_excel_files(input_root))
    results: List[UploadResult] = []

    for file_path in files:
        source_mime = get_upload_mime_type(file_path)
        target_mime = (
            "application/vnd.google-apps.spreadsheet"
            if convert_to_sheets
            else source_mime
        )
        target_name = file_path.stem if convert_to_sheets else file_path.name

        metadata = {"name": target_name}
        if drive_folder_id:
            metadata["parents"] = [drive_folder_id]
        if convert_to_sheets:
            metadata["mimeType"] = "application/vnd.google-apps.spreadsheet"

        try:
            media = MediaFileUpload(str(file_path), mimetype=source_mime, resumable=True)
            uploaded = (
                service.files()
                .create(
                    body=metadata,
                    media_body=media,
                    fields="id,name,mimeType,webViewLink",
                )
                .execute()
            )
            results.append(
                UploadResult(
                    local_path=str(file_path),
                    file_name=uploaded.get("name", target_name),
                    source_mime_type=source_mime,
                    target_mime_type=uploaded.get("mimeType", target_mime),
                    drive_file_id=uploaded.get("id", ""),
                    drive_web_view_link=uploaded.get("webViewLink", ""),
                    status="success",
                    error="",
                )
            )
        except Exception as exc:
            results.append(
                UploadResult(
                    local_path=str(file_path),
                    file_name=target_name,
                    source_mime_type=source_mime,
                    target_mime_type=target_mime,
                    drive_file_id="",
                    drive_web_view_link="",
                    status="failed",
                    error=str(exc),
                )
            )

    success_count = sum(1 for r in results if r.status == "success")
    failure_count = len(results) - success_count

    report = UploadReport(
        generated_utc=datetime.now(tz=timezone.utc).isoformat(),
        input_root=str(input_root),
        output_root=str(output_dir),
        file_count=len(results),
        success_count=success_count,
        failure_count=failure_count,
        converted_to_sheets=convert_to_sheets,
        drive_folder_id=drive_folder_id or "",
        files=results,
    )

    output_dir.mkdir(parents=True, exist_ok=True)
    write_upload_json(report, output_dir)
    write_upload_csv(report, output_dir)
    return report


def extract_form_controls_from_xlsm(file_path: Path) -> List[FormControl]:
    """Extract form control buttons from an xlsm file by parsing VML drawings."""
    controls: List[FormControl] = []
    if file_path.suffix.lower() != ".xlsm":
        return controls

    try:
        with zipfile.ZipFile(file_path, "r") as zf:
            # Map drawing relationships to sheet names via rels
            sheet_for_drawing: dict[str, str] = {}
            # Read workbook to get sheet names
            try:
                wb_xml = zf.read("xl/workbook.xml").decode("utf-8")
            except KeyError:
                return controls

            sheet_names: list[str] = re.findall(r'<sheet[^>]+name="([^"]+)"', wb_xml)

            # Map rId -> sheet index from workbook rels
            try:
                wb_rels = zf.read("xl/_rels/workbook.xml.rels").decode("utf-8")
            except KeyError:
                wb_rels = ""

            rid_to_target: dict[str, str] = {}
            for m in re.finditer(r'<Relationship[^>]+Id="([^"]+)"[^>]+Target="([^"]+)"', wb_rels):
                rid_to_target[m.group(1)] = m.group(2)

            # Map sheet rIds from workbook.xml
            sheet_rids = re.findall(r'<sheet[^>]+r:id="([^"]+)"', wb_xml)
            sheet_target_to_name: dict[str, str] = {}
            for i, rid in enumerate(sheet_rids):
                target = rid_to_target.get(rid, "")
                if target:
                    # Normalize: "worksheets/sheet1.xml" -> "sheet1"
                    sheet_key = target.replace("worksheets/", "").replace(".xml", "")
                    if i < len(sheet_names):
                        sheet_target_to_name[sheet_key] = sheet_names[i]

            # Find sheet -> vmlDrawing mapping from sheet rels
            for name in zf.namelist():
                m = re.match(r"xl/worksheets/_rels/(sheet\d+)\.xml\.rels$", name)
                if not m:
                    continue
                sheet_key = m.group(1)
                sheet_name = sheet_target_to_name.get(sheet_key, sheet_key)
                rels_xml = zf.read(name).decode("utf-8")
                for rel_match in re.finditer(
                    r'<Relationship[^>]+Target="([^"]*vmlDrawing[^"]*)"', rels_xml
                ):
                    vml_target = rel_match.group(1)
                    # Normalize path: ../drawings/vmlDrawing1.vml -> xl/drawings/vmlDrawing1.vml
                    if vml_target.startswith(".."):
                        vml_path = "xl/" + vml_target.lstrip("./")
                    elif not vml_target.startswith("xl/"):
                        vml_path = "xl/drawings/" + vml_target
                    else:
                        vml_path = vml_target
                    sheet_for_drawing[vml_path] = sheet_name

            # Parse each VML drawing file
            for vml_path, sheet_name in sheet_for_drawing.items():
                try:
                    raw = zf.read(vml_path)
                except KeyError:
                    continue
                # Try UTF-8 first, then Shift-JIS (common for Japanese Excel)
                for enc in ("utf-8", "cp932", "latin-1"):
                    try:
                        vml_data = raw.decode(enc)
                        break
                    except (UnicodeDecodeError, LookupError):
                        continue
                else:
                    vml_data = raw.decode("latin-1")

                # Split into shape blocks
                for shape_block in re.finditer(
                    r"<v:shape\b[^>]*>(.*?)</v:shape>", vml_data, re.DOTALL
                ):
                    block = shape_block.group(1)
                    # Check if this is a Button type
                    obj_type_m = re.search(
                        r'<x:ClientData\s+ObjectType="(\w+)"', block
                    )
                    if not obj_type_m or obj_type_m.group(1) != "Button":
                        continue

                    # Extract label (text content)
                    label = ""
                    text_m = re.search(r"<v:textbox[^>]*>(.*?)</v:textbox>", block, re.DOTALL)
                    if text_m:
                        # Strip HTML/XML tags to get plain text
                        raw_label = re.sub(r"<[^>]+>", "", text_m.group(1))
                        # Normalize whitespace (newlines, multiple spaces)
                        label = " ".join(raw_label.split())

                    # Extract macro name from FmlaMacro
                    macro = ""
                    macro_m = re.search(r"<x:FmlaMacro>(.*?)</x:FmlaMacro>", block, re.DOTALL)
                    if macro_m:
                        macro = macro_m.group(1).strip()
                        # Remove workbook index prefix like [0]!
                        macro = re.sub(r"^\[\d+\]!", "", macro)

                    # Extract anchor coordinates
                    anchor = ""
                    anchor_m = re.search(r"<x:Anchor>(.*?)</x:Anchor>", block, re.DOTALL)
                    if anchor_m:
                        coords = [c.strip() for c in anchor_m.group(1).split(",")]
                        if len(coords) >= 8:
                            # coords: col_start, offset, row_start, offset, col_end, offset, row_end, offset
                            col_start = int(coords[0])
                            row_start = int(coords[2])
                            col_end = int(coords[4])
                            row_end = int(coords[6])

                            def col_letter(n: int) -> str:
                                result = ""
                                while n >= 0:
                                    result = chr(ord("A") + n % 26) + result
                                    n = n // 26 - 1
                                return result

                            anchor = f"{col_letter(col_start)}{row_start + 1}:{col_letter(col_end)}{row_end + 1}"

                    control_name = label or "Button"
                    controls.append(
                        FormControl(
                            source_file=str(file_path),
                            sheet_name=sheet_name,
                            control_name=control_name,
                            control_type="Button",
                            label=label,
                            macro=macro,
                            anchor=anchor,
                        )
                    )
    except (zipfile.BadZipFile, OSError):
        pass

    return controls


def _extract_sheet_data(sheet, index: int) -> dict:
    """Extract all non-empty cells, formulas, and data validations from a sheet."""
    cells: dict[str, dict] = {}
    formulas: list[dict] = []

    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            coord = cell.coordinate
            value = cell.value

            if isinstance(value, str) and value.startswith("="):
                cell_type = "formula"
                formulas.append({"cell": coord, "formula": value})
            elif isinstance(value, bool):
                cell_type = "boolean"
            elif isinstance(value, (int, float)):
                cell_type = "number"
            elif isinstance(value, datetime):
                value = value.isoformat()
                cell_type = "datetime"
            else:
                cell_type = "string"
                value = str(value)

            cells[coord] = {"value": value, "type": cell_type}

    data_validations: list[dict] = []
    if hasattr(sheet, "data_validations") and sheet.data_validations:
        for dv in sheet.data_validations.dataValidation:
            dv_dict: dict = {"type": dv.type or "", "sqref": str(dv.sqref)}
            if dv.formula1:
                dv_dict["formula1"] = dv.formula1
            if dv.formula2:
                dv_dict["formula2"] = dv.formula2
            if dv.allow_blank is not None:
                dv_dict["allow_blank"] = dv.allow_blank
            data_validations.append(dv_dict)

    return {
        "name": sheet.title,
        "index": index,
        "row_count": sheet.max_row or 0,
        "col_count": sheet.max_column or 0,
        "cells": cells,
        "formulas": formulas,
        "data_validations": data_validations,
    }


def _extract_named_ranges(workbook) -> list[dict]:
    """Extract named ranges from the workbook."""
    named_ranges: list[dict] = []
    for dn in workbook.defined_names.values():
        scope = "global"
        if dn.localSheetId is not None:
            try:
                scope = workbook.sheetnames[dn.localSheetId]
            except IndexError:
                scope = f"sheet_{dn.localSheetId}"
        named_ranges.append({
            "name": dn.name,
            "value": dn.attr_text,
            "scope": scope,
        })
    return named_ranges


_INTER_SHEET_RE = re.compile(r"(?:'([^']+)'|([A-Za-z_]\w*))\!")


def _extract_inter_sheet_references(sheets_data: list[dict], sheet_names: list[str]) -> list[dict]:
    """Detect cross-sheet references in formulas."""
    refs: list[dict] = []
    sheet_name_set = set(sheet_names)

    for sd in sheets_data:
        for f in sd["formulas"]:
            targets: list[str] = []
            for m in _INTER_SHEET_RE.finditer(f["formula"]):
                target = m.group(1) or m.group(2)
                if target in sheet_name_set and target != sd["name"]:
                    targets.append(target)
            if targets:
                refs.append({
                    "source_sheet": sd["name"],
                    "source_cell": f["cell"],
                    "formula": f["formula"],
                    "target_sheets": sorted(set(targets)),
                })
    return refs


_VBA_PUBLIC_FUNC_RE = re.compile(
    r"^\s*Public\s+Function\s+(\w+)\s*\(([^)]*)\)",
    re.MULTILINE | re.IGNORECASE,
)


def _extract_vba_public_functions(modules_data: list[dict]) -> dict[str, dict]:
    """Extract Public Function definitions from VBA Standard modules.

    Returns {"FUNCNAME": {"module": "Module1", "signature": "..."}} (upper-cased keys).
    """
    udf_map: dict[str, dict] = {}
    for mod in modules_data:
        if mod.get("module_type") != "Standard":
            continue
        # Read .bas file content
        bas_path = mod.get("output_path", "")
        if not bas_path:
            continue
        try:
            code = Path(bas_path).read_text(encoding="utf-8", errors="replace")
        except OSError:
            continue
        for m in _VBA_PUBLIC_FUNC_RE.finditer(code):
            name = m.group(1)
            sig = f"{name}({m.group(2).strip()})"
            udf_map[name.upper()] = {"module": mod.get("module_name", ""), "signature": sig}
    return udf_map


def _analyze_formula_compatibility(
    sheets_data: list[dict], vba_udf_map: dict[str, dict]
) -> dict:
    """Scan all formulas for incompatible Excel functions and VBA UDF calls.

    Returns dict with incompatible_formulas, udf_calls, and summary.
    """
    incompatible: list[dict] = []
    udf_calls: list[dict] = []

    for sd in sheets_data:
        for f in sd.get("formulas", []):
            upper = f["formula"].upper()
            # Check incompatible built-in functions
            for pattern, category in INCOMPATIBLE_FUNCTIONS.items():
                if pattern in upper:
                    func_name = pattern.rstrip("(")
                    incompatible.append({
                        "sheet": sd["name"],
                        "cell": f["cell"],
                        "formula": f["formula"],
                        "function": func_name,
                        "category": category,
                    })
            # Check VBA UDF usage in formulas
            for udf_name, udf_info in vba_udf_map.items():
                if udf_name + "(" in upper:
                    udf_calls.append({
                        "sheet": sd["name"],
                        "cell": f["cell"],
                        "formula": f["formula"],
                        "udf_name": udf_name,
                        "module": udf_info["module"],
                        "signature": udf_info["signature"],
                    })

    missing = [i for i in incompatible if i["category"] == "missing"]
    partial = [i for i in incompatible if i["category"] == "partial"]
    check = [i for i in incompatible if i["category"] == "check"]

    return {
        "incompatible_formulas": incompatible,
        "udf_calls": udf_calls,
        "summary": {
            "total_incompatible": len(incompatible),
            "missing_count": len(missing),
            "partial_count": len(partial),
            "check_count": len(check),
            "total_udf_calls": len(udf_calls),
            "unique_missing_functions": sorted({i["function"] for i in missing}),
            "unique_udf_names": sorted({u["udf_name"] for u in udf_calls}),
        },
    }


def extract_workbook_context(
    file_path: Path,
    modules: Optional[List[VbaModule]] = None,
    controls: Optional[List[FormControl]] = None,
) -> dict:
    """Extract comprehensive workbook context for VBA conversion."""
    workbook = load_workbook(filename=file_path, data_only=False, keep_vba=True)
    try:
        sheets_data: list[dict] = []
        for i, sheet in enumerate(workbook.worksheets):
            sheets_data.append(_extract_sheet_data(sheet, i))

        named_ranges = _extract_named_ranges(workbook)
        inter_sheet_refs = _extract_inter_sheet_references(
            sheets_data, workbook.sheetnames
        )

        # Use provided controls/modules or extract fresh
        controls_data = (
            [asdict(c) for c in controls]
            if controls is not None
            else [asdict(c) for c in extract_form_controls_from_xlsm(file_path)]
        )
        modules_data = (
            [asdict(m) for m in modules] if modules is not None else []
        )

        total_non_empty = sum(len(sd["cells"]) for sd in sheets_data)
        total_formulas = sum(len(sd["formulas"]) for sd in sheets_data)
        total_dv = sum(len(sd["data_validations"]) for sd in sheets_data)

        # Analyze formula compatibility (incompatible functions + VBA UDFs)
        vba_udf_map = _extract_vba_public_functions(modules_data)
        func_compat = _analyze_formula_compatibility(sheets_data, vba_udf_map)

        return {
            "source_file": str(file_path),
            "generated_utc": datetime.now(tz=timezone.utc).isoformat(),
            "modules": modules_data,
            "sheets": sheets_data,
            "named_ranges": named_ranges,
            "inter_sheet_references": inter_sheet_refs,
            "controls": controls_data,
            "function_compatibility": func_compat,
            "summary": {
                "sheet_count": len(workbook.sheetnames),
                "total_non_empty_cells": total_non_empty,
                "total_formulas": total_formulas,
                "total_data_validations": total_dv,
                "total_named_ranges": len(named_ranges),
                "total_inter_sheet_refs": len(inter_sheet_refs),
                "total_controls": len(controls_data),
                "total_modules": len(modules_data),
                "total_incompatible_functions": func_compat["summary"]["total_incompatible"],
                "total_udf_calls": func_compat["summary"]["total_udf_calls"],
            },
        }
    finally:
        workbook.close()


def write_workbook_context(context: dict, output_dir: Path) -> Path:
    """Write workbook_context.json to the output directory."""
    output_path = output_dir / "workbook_context.json"
    output_path.write_text(
        json.dumps(context, ensure_ascii=False, indent=2, default=str),
        encoding="utf-8",
    )
    return output_path


def find_xlsm_files(root: Path) -> Iterable[Path]:
    if root.is_file():
        if root.suffix.lower() == ".xlsm":
            yield root
        return
    for file_path in root.rglob("*.xlsm"):
        if file_path.is_file():
            yield file_path


def classify_vba_module(module_name: str, code: str) -> str:
    lower_name = module_name.lower()
    if lower_name == "thisworkbook":
        return "ThisWorkbook"
    if lower_name.startswith("sheet") or "Attribute VB_Base = \"0{00020820" in code:
        return "Sheet"
    if "Attribute VB_Base = \"0{" in code:
        return "UserForm"
    if "Attribute VB_Creatable = True" in code or "VERSION 1.0 CLASS" in code:
        return "Class"
    return "Standard"


def extract_vba_from_file(
    file_path: Path, output_dir: Path
) -> List[VbaModule]:
    from oletools.olevba import VBA_Parser

    parser = VBA_Parser(str(file_path))
    modules: List[VbaModule] = []

    if not parser.detect_vba_macros():
        parser.close()
        return modules

    file_output_dir = output_dir / file_path.stem
    file_output_dir.mkdir(parents=True, exist_ok=True)

    for _, _, vba_filename, vba_code in parser.extract_macros():
        if not vba_code or not vba_code.strip():
            continue

        module_name = Path(vba_filename).stem if vba_filename else "Unknown"
        module_type = classify_vba_module(module_name, vba_code)
        code_lines = len(vba_code.splitlines())

        bas_path = file_output_dir / f"{module_name}.bas"
        bas_path.write_text(vba_code, encoding="utf-8")

        modules.append(
            VbaModule(
                source_file=str(file_path),
                module_name=module_name,
                module_type=module_type,
                code_lines=code_lines,
                output_path=str(bas_path),
            )
        )

    parser.close()
    return modules


def write_extract_json(report: ExtractReport, output_dir: Path) -> Path:
    output_path = output_dir / "extract_report.json"
    output_path.write_text(
        json.dumps(
            {
                "generated_utc": report.generated_utc,
                "input_root": report.input_root,
                "file_count": report.file_count,
                "module_count": report.module_count,
                "modules": [asdict(m) for m in report.modules],
                "controls": [asdict(c) for c in report.controls],
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    return output_path


def write_extract_csv(report: ExtractReport, output_dir: Path) -> Path:
    output_path = output_dir / "extract_report.csv"
    fieldnames = ["source_file", "module_name", "module_type", "code_lines", "output_path"]
    with output_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for m in report.modules:
            writer.writerow(asdict(m))
    return output_path


def run_extract(input_root: Path, output_dir: Path) -> ExtractReport:
    files = list(find_xlsm_files(input_root))
    all_modules: List[VbaModule] = []
    all_controls: List[FormControl] = []
    for file_path in files:
        modules: List[VbaModule] = []
        ctrls: List[FormControl] = []

        try:
            modules = extract_vba_from_file(file_path, output_dir)
            all_modules.extend(modules)
            print(f"  {file_path.name}: {len(modules)} modules")
        except Exception as exc:
            print(f"  {file_path.name}: ERROR - {exc}")

        try:
            ctrls = extract_form_controls_from_xlsm(file_path)
            all_controls.extend(ctrls)
            if ctrls:
                print(f"  {file_path.name}: {len(ctrls)} form controls")
        except Exception as exc:
            print(f"  {file_path.name}: form control extraction error - {exc}")

        # Extract workbook context (sheets, named ranges, data validations, etc.)
        try:
            context = extract_workbook_context(file_path, modules, ctrls)
            file_output_dir = output_dir / file_path.stem
            file_output_dir.mkdir(parents=True, exist_ok=True)
            write_workbook_context(context, file_output_dir)
            summary = context["summary"]
            print(
                f"  {file_path.name}: context extracted "
                f"(sheets={summary['sheet_count']}, "
                f"cells={summary['total_non_empty_cells']}, "
                f"formulas={summary['total_formulas']}, "
                f"validations={summary['total_data_validations']}, "
                f"named_ranges={summary['total_named_ranges']})"
            )
            # Warn about incompatible functions and UDF usage
            n_incompat = summary.get("total_incompatible_functions", 0)
            n_udf = summary.get("total_udf_calls", 0)
            if n_incompat:
                fc_summary = context["function_compatibility"]["summary"]
                funcs = ", ".join(fc_summary.get("unique_missing_functions", []))
                print(
                    f"  WARNING: {n_incompat} incompatible function call(s) detected"
                    f" ({funcs})" if funcs else ""
                )
            if n_udf:
                fc_summary = context["function_compatibility"]["summary"]
                udfs = ", ".join(fc_summary.get("unique_udf_names", []))
                print(
                    f"  WARNING: {n_udf} VBA UDF call(s) in formulas"
                    f" ({udfs})" if udfs else ""
                )
        except Exception as exc:
            print(f"  {file_path.name}: context extraction error - {exc}")

    report = ExtractReport(
        generated_utc=datetime.now(tz=timezone.utc).isoformat(),
        input_root=str(input_root),
        file_count=len(files),
        module_count=len(all_modules),
        modules=all_modules,
        controls=all_controls,
    )

    output_dir.mkdir(parents=True, exist_ok=True)
    write_extract_json(report, output_dir)
    write_extract_csv(report, output_dir)
    return report


VBA_TO_GAS_SYSTEM_PROMPT = """\
You are a VBA-to-Google Apps Script converter.
Convert the given VBA module to Google Apps Script (JavaScript).

Key conversion rules:
- ThisWorkbook → SpreadsheetApp.getActiveSpreadsheet()
- ActiveSheet / Worksheets("name") → getSheetByName("name")
- Range("A1") / Cells(row, col) → getRange() / getRange(row, col)
- .Value → .getValue() / .setValue()
- MsgBox → Browser.msgBox()
- VBA Sub/Function → JavaScript function
- For...Next → for loop
- Dim → let/const
- VBA string concatenation (&) → JavaScript (+)
- VBA comments (') → JavaScript comments (//)
- On Error → try/catch
- Nothing → null

Excel form control buttons do not exist in Google Sheets.
When button information is provided, generate an onOpen() trigger that creates
a custom menu named "Macros" using SpreadsheetApp.getUi().createMenu("Macros").
Add each button as a menu item via .addItem(buttonLabel, functionName).

When workbook context information is provided (sheet data, named ranges,
data validations, cross-sheet references), use it to:
- Correctly reference sheet names and cell ranges in the converted code.
- Preserve data validation logic (e.g. list dropdowns) using
  SpreadsheetApp Range.setDataValidation().
- Map named ranges to NamedRange or getRangeByName() calls.
- Ensure cross-sheet references use getSheetByName() with the correct names.

Function compatibility handling:
When function compatibility information is provided, follow these rules:
- VBA UDFs (user-defined functions) used in cell formulas: convert to Apps Script
  custom functions with the /** @customfunction */ JSDoc tag so they work in cells.
- "missing" functions (no Google Sheets equivalent): generate a setupFormulas()
  helper that uses Range.setFormula() or Range.setValue() to provide a workaround.
  Add a comment explaining the original Excel function and the limitation.
- "partial" functions: convert directly but add a // WARNING comment noting the
  behavioural difference from Excel.
- "check" functions: convert directly with a // NOTE comment to verify behaviour.

Output ONLY the converted code. No explanations, no markdown fences.
"""


def extract_code_block(text: str) -> str:
    """Extract code from markdown fenced block if present, else return full text."""
    m = re.search(r"```(?:javascript|js|gs|gas)?\s*\n(.*?)```", text, re.DOTALL)
    if m:
        return m.group(1).strip()
    m = re.search(r"```\s*\n(.*?)```", text, re.DOTALL)
    if m:
        return m.group(1).strip()
    return text.strip()


def convert_vba_module(
    vba_code: str, module_name: str, module_type: str, client, model: str,
    controls_context: str = "",
    workbook_context: str = "",
    func_compat_context: str = "",
) -> str:
    """Call Anthropic Messages API to convert a single VBA module to Apps Script."""
    user_msg = f"Module: {module_name} (type: {module_type})\n\n{vba_code}"
    if controls_context:
        user_msg += f"\n\n{controls_context}"
    if workbook_context:
        user_msg += f"\n\n{workbook_context}"
    if func_compat_context:
        user_msg += f"\n\n{func_compat_context}"
    response = client.messages.create(
        model=model,
        max_tokens=4096,
        system=VBA_TO_GAS_SYSTEM_PROMPT,
        messages=[{"role": "user", "content": user_msg}],
    )
    raw = response.content[0].text
    return extract_code_block(raw)


def write_convert_json(report: ConvertReport, output_dir: Path) -> Path:
    output_path = output_dir / "convert_report.json"
    output_path.write_text(
        json.dumps(
            {
                "generated_utc": report.generated_utc,
                "input_root": report.input_root,
                "total": report.total,
                "success": report.success,
                "failed": report.failed,
                "results": [asdict(r) for r in report.results],
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    return output_path


def write_convert_csv(report: ConvertReport, output_dir: Path) -> Path:
    output_path = output_dir / "convert_report.csv"
    fieldnames = ["source_bas", "module_name", "module_type", "output_gs", "status", "error"]
    with output_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for r in report.results:
            writer.writerow(asdict(r))
    return output_path


def run_convert(
    input_dir: Path, output_dir: Path, api_key: Optional[str], model: str
) -> ConvertReport:
    from anthropic import Anthropic

    resolved_key = api_key or os.environ.get("ANTHROPIC_API_KEY", "")
    if not resolved_key:
        raise SystemExit("No API key provided. Use --api-key or set ANTHROPIC_API_KEY.")

    client = Anthropic(api_key=resolved_key)
    output_dir.mkdir(parents=True, exist_ok=True)

    bas_files = sorted(input_dir.rglob("*.bas"))
    if not bas_files:
        print(f"No .bas files found under {input_dir}")

    # Load workbook_context.json from each workbook subdirectory.
    # This single file contains modules, controls, sheets, named ranges, etc.
    wb_data_map: dict[str, dict] = {}
    for sub in sorted(input_dir.iterdir()):
        ctx_path = sub / "workbook_context.json" if sub.is_dir() else None
        if ctx_path and ctx_path.exists():
            try:
                wb_data_map[sub.name] = json.loads(
                    ctx_path.read_text(encoding="utf-8")
                )
            except Exception:
                pass

    results: List[ConvertResult] = []
    for bas_path in bas_files:
        module_name = bas_path.stem
        # Preserve subdirectory structure (e.g. workbook_name/)
        rel = bas_path.parent.relative_to(input_dir)
        gs_dir = output_dir / rel
        gs_dir.mkdir(parents=True, exist_ok=True)
        gs_path = gs_dir / f"{module_name}.gs"

        # Look up module_type and context from workbook_context.json
        wb_ctx_data = wb_data_map.get(str(rel), {})

        module_type = "Standard"
        for m in wb_ctx_data.get("modules", []):
            if m.get("module_name") == module_name:
                module_type = m.get("module_type", "Standard")
                break

        # Build controls context for ThisWorkbook modules
        controls_context = ""
        if module_type == "ThisWorkbook":
            controls_list = wb_ctx_data.get("controls", [])
            if controls_list:
                lines = ["Excel form control buttons found in this workbook:"]
                for c in controls_list:
                    lines.append(
                        f"- Button label=\"{c.get('label', '')}\" macro=\"{c.get('macro', '')}\" "
                        f"sheet=\"{c.get('sheet_name', '')}\" anchor=\"{c.get('anchor', '')}\""
                    )
                lines.append(
                    "\nGenerate an onOpen() function that creates a 'Macros' custom menu "
                    "with menu items for each button above."
                )
                controls_context = "\n".join(lines)

        # Build workbook context string (sheets, formulas, validations, etc.)
        wb_ctx = ""
        if wb_ctx_data:
            parts = ["Workbook context information:"]
            for sd in wb_ctx_data.get("sheets", []):
                parts.append(f"\nSheet \"{sd['name']}\" ({sd.get('row_count', 0)} rows x {sd.get('col_count', 0)} cols):")
                for f in sd.get("formulas", []):
                    parts.append(f"  {f['cell']}: {f['formula']}")
                for dv in sd.get("data_validations", []):
                    parts.append(f"  Validation [{dv.get('type', '')}] at {dv.get('sqref', '')}: {dv.get('formula1', '')}")
            for nr in wb_ctx_data.get("named_ranges", []):
                parts.append(f"Named range: {nr['name']} = {nr['value']} (scope: {nr['scope']})")
            for ref in wb_ctx_data.get("inter_sheet_references", []):
                parts.append(f"Cross-sheet ref: {ref['source_sheet']}!{ref['source_cell']} -> {ref['formula']}")
            wb_ctx = "\n".join(parts)

        # Build function compatibility context
        fc_ctx = ""
        func_compat = wb_ctx_data.get("function_compatibility", {})
        if func_compat:
            fc_parts = ["Function compatibility issues in this workbook:"]
            for item in func_compat.get("incompatible_formulas", []):
                fc_parts.append(
                    f"- [{item['category']}] {item['function']}() in "
                    f"{item['sheet']}!{item['cell']}: {item['formula']}"
                )
            for item in func_compat.get("udf_calls", []):
                fc_parts.append(
                    f"- [UDF] {item['udf_name']}() from {item['module']} "
                    f"in {item['sheet']}!{item['cell']}: {item['formula']}"
                )
            if len(fc_parts) > 1:
                fc_ctx = "\n".join(fc_parts)

        print(f"  Converting {rel / bas_path.name} ...", end=" ", flush=True)
        try:
            vba_code = bas_path.read_text(encoding="utf-8", errors="replace")
            gs_code = convert_vba_module(
                vba_code, module_name, module_type, client, model,
                controls_context=controls_context, workbook_context=wb_ctx,
                func_compat_context=fc_ctx,
            )
            gs_path.write_text(gs_code, encoding="utf-8")
            results.append(
                ConvertResult(
                    source_bas=str(bas_path),
                    module_name=module_name,
                    module_type=module_type,
                    output_gs=str(gs_path),
                    status="success",
                    error="",
                )
            )
            print("OK")
        except Exception as exc:
            results.append(
                ConvertResult(
                    source_bas=str(bas_path),
                    module_name=module_name,
                    module_type=module_type,
                    output_gs="",
                    status="failed",
                    error=str(exc),
                )
            )
            print(f"FAILED - {exc}")

    # Warn about button controls converted to custom menus
    has_controls = any(
        wb.get("controls") for wb in wb_data_map.values()
    )
    if has_controls:
        print(
            "\nNote: Excel form control buttons have been converted to "
            "onOpen() custom menus.\n"
            "Google Sheets does not support programmatic button creation.\n"
            "To add a button on the sheet, manually insert a drawing "
            "(Insert > Drawing) and assign the script."
        )

    # Warn about incompatible functions and UDFs
    for wb_name, wb in wb_data_map.items():
        fc = wb.get("function_compatibility", {})
        fc_sum = fc.get("summary", {})
        n_incompat = fc_sum.get("total_incompatible", 0)
        n_udf = fc_sum.get("total_udf_calls", 0)
        if n_incompat:
            missing = fc_sum.get("unique_missing_functions", [])
            print(
                f"\nWARNING [{wb_name}]: {n_incompat} incompatible Excel function(s) detected."
            )
            if missing:
                print(f"  Missing (no Sheets equivalent): {', '.join(missing)}")
            print(
                "  These have been handled via setupFormulas() or comments in the output."
            )
        if n_udf:
            udfs = fc_sum.get("unique_udf_names", [])
            print(
                f"\nWARNING [{wb_name}]: {n_udf} VBA UDF call(s) found in cell formulas."
            )
            if udfs:
                print(f"  UDFs: {', '.join(udfs)}")
            print(
                "  These have been converted to @customfunction Apps Script functions."
            )

    success = sum(1 for r in results if r.status == "success")
    report = ConvertReport(
        generated_utc=datetime.now(tz=timezone.utc).isoformat(),
        input_root=str(input_dir),
        total=len(results),
        success=success,
        failed=len(results) - success,
        results=results,
    )

    write_convert_json(report, output_dir)
    write_convert_csv(report, output_dir)
    return report


def load_script_service(credentials_path: Path, token_path: Path):
    from google.auth.exceptions import RefreshError
    from google.auth.transport.requests import Request
    from google.oauth2.credentials import Credentials
    from google_auth_oauthlib.flow import InstalledAppFlow
    from googleapiclient.discovery import build

    creds = None
    if token_path.exists():
        # Check actual scopes stored in token file before loading
        try:
            token_data = json.loads(token_path.read_text(encoding="utf-8"))
            file_scopes = set(token_data.get("scopes", []))
            if not set(SCRIPT_SCOPES).issubset(file_scopes):
                # Token missing required scopes - delete and re-authenticate
                token_path.unlink()
            else:
                creds = Credentials.from_authorized_user_file(
                    str(token_path), SCRIPT_SCOPES
                )
        except Exception:
            pass

    if not creds or not creds.valid:
        refreshed = False
        if creds and creds.expired and creds.refresh_token:
            try:
                creds.refresh(Request())
                refreshed = True
            except RefreshError:
                pass
        if not refreshed:
            flow = InstalledAppFlow.from_client_secrets_file(str(credentials_path), SCRIPT_SCOPES)
            creds = flow.run_local_server(port=0)
        token_path.parent.mkdir(parents=True, exist_ok=True)
        token_path.write_text(creds.to_json(), encoding="utf-8")

    return build("script", "v1", credentials=creds, cache_discovery=False)


def write_deploy_json(report: DeployReport, output_dir: Path) -> Path:
    output_path = output_dir / "deploy_report.json"
    output_path.write_text(
        json.dumps(asdict(report), ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return output_path


def run_deploy(
    input_dir: Path,
    spreadsheet_id: str,
    credentials_path: Path,
    token_path: Path,
    script_id: Optional[str] = None,
) -> DeployReport:
    service = load_script_service(credentials_path, token_path)

    gs_files = sorted(input_dir.rglob("*.gs"))
    if not gs_files:
        raise SystemExit(f"No .gs files found under {input_dir}")

    print(f"Deploying to spreadsheet: {spreadsheet_id}")

    if script_id:
        print(f"  Using existing script project: {script_id}")
    else:
        create_body = {"title": "ExcelMigration", "parentId": spreadsheet_id}
        created = service.projects().create(body=create_body).execute()
        script_id = created["scriptId"]
        print(f"  Created script project: {script_id}")

    manifest = {
        "name": "appsscript",
        "type": "JSON",
        "source": json.dumps(
            {
                "timeZone": "Asia/Tokyo",
                "dependencies": {},
                "exceptionLogging": "STACKDRIVER",
                "runtimeVersion": "V8",
                "oauthScopes": [
                    "https://www.googleapis.com/auth/spreadsheets.currentonly"
                ],
            },
            ensure_ascii=False,
        ),
    }

    files_payload = [manifest]
    deployed_names: List[str] = []
    print(f"  Pushing {len(gs_files)} files...")
    for gs_path in gs_files:
        name = gs_path.stem
        source = gs_path.read_text(encoding="utf-8")
        files_payload.append({"name": name, "type": "SERVER_JS", "source": source})
        deployed_names.append(gs_path.name)

    body = {"scriptId": script_id, "files": files_payload}
    service.projects().updateContent(scriptId=script_id, body=body).execute()

    for name in deployed_names:
        print(f"  {name} ... OK")
    print(f"Deploy complete: {len(deployed_names)} files pushed")
    print(f"Script editor: https://script.google.com/d/{script_id}/edit")

    report = DeployReport(
        generated_utc=datetime.now(tz=timezone.utc).isoformat(),
        spreadsheet_id=spreadsheet_id,
        script_id=script_id,
        file_count=len(deployed_names),
        files_deployed=deployed_names,
    )

    input_dir.parent.mkdir(parents=True, exist_ok=True)
    write_deploy_json(report, input_dir)
    return report


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Excel migration inventory CLI")
    subparsers = parser.add_subparsers(dest="command", required=True)

    scan_parser = subparsers.add_parser("scan", help="Scan and analyze Excel files")
    scan_parser.add_argument("--input", required=True, help="Root folder to scan")
    scan_parser.add_argument("--output", required=True, help="Folder for report outputs")
    scan_parser.add_argument(
        "--group-by",
        choices=["prefix", "subfolder", "none"],
        default="none",
        help="Group files by filename prefix or subfolder (default: none)",
    )

    extract_parser = subparsers.add_parser("extract", help="Extract VBA modules from .xlsm files")
    extract_parser.add_argument("--input", required=True, help=".xlsm file or folder to scan")
    extract_parser.add_argument("--output", required=True, help="Folder for extracted .bas files and reports")

    convert_parser = subparsers.add_parser("convert", help="Convert VBA modules to Google Apps Script via Claude API")
    convert_parser.add_argument("--input", required=True, help="Folder with extracted .bas files")
    convert_parser.add_argument("--output", required=True, help="Folder for converted .gs files and reports")
    convert_parser.add_argument("--api-key", default="", help="Anthropic API key (default: env ANTHROPIC_API_KEY)")
    convert_parser.add_argument("--model", default="claude-sonnet-4-6", help="Model to use (default: claude-sonnet-4-6)")

    upload_parser = subparsers.add_parser("upload", help="Upload Excel files to Google Drive")
    upload_parser.add_argument("--input", required=True, help="Root folder to scan and upload")
    upload_parser.add_argument("--output", required=True, help="Folder for upload reports")
    upload_parser.add_argument(
        "--credentials",
        default="credentials/client_secret.json",
        help="OAuth client secret JSON from Google Cloud",
    )
    upload_parser.add_argument(
        "--token",
        default="credentials/token.json",
        help="Path to OAuth token cache file",
    )
    upload_parser.add_argument(
        "--drive-folder-id",
        default="",
        help="Drive folder ID for uploaded files (optional)",
    )
    upload_parser.add_argument(
        "--convert-to-sheets",
        action="store_true",
        help="Convert uploaded Excel files into Google Sheets format",
    )

    deploy_parser = subparsers.add_parser("deploy", help="Deploy .gs files to a Google Spreadsheet via Apps Script API")
    deploy_parser.add_argument("--input", required=True, help="Folder containing .gs files")
    deploy_parser.add_argument("--spreadsheet-id", required=True, help="Target Google Spreadsheet ID")
    deploy_parser.add_argument("--script-id", default="", help="Existing Apps Script project ID (optional; creates new if omitted)")
    deploy_parser.add_argument(
        "--credentials",
        default="credentials/client_secret.json",
        help="OAuth client secret JSON from Google Cloud",
    )
    deploy_parser.add_argument(
        "--token",
        default="credentials/token.json",
        help="Path to OAuth token cache file",
    )

    migrate_parser = subparsers.add_parser(
        "migrate", help="End-to-end: upload xlsm to Sheets and deploy .gs files"
    )
    migrate_parser.add_argument("--xlsm", required=True, help="Source .xlsm file")
    migrate_parser.add_argument("--gs-dir", required=True, help="Folder with converted .gs files")
    migrate_parser.add_argument("--output", required=True, help="Folder for migration reports")
    migrate_parser.add_argument(
        "--credentials",
        default="credentials/client_secret.json",
        help="OAuth client secret JSON from Google Cloud",
    )
    migrate_parser.add_argument(
        "--token",
        default="credentials/token.json",
        help="Path to OAuth token cache file",
    )

    return parser


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()

    if args.command == "scan":
        input_root = Path(args.input).expanduser().resolve()
        output_dir = Path(args.output).expanduser().resolve()
        group_by = args.group_by
        report = run_scan(input_root, output_dir, group_by=group_by)
        print(f"Scanned files: {report.file_count}")
        print(f"JSON report: {output_dir / 'report.json'}")
        print(f"CSV report: {output_dir / 'report.csv'}")
        if report.groups:
            print(f"Group summary CSV: {output_dir / 'group_summary.csv'}")
            print(f"Groups: {len(report.groups)}")
            for g in report.groups:
                print(f"  {g.group_name}: {g.file_count} files, {g.migration_difficulty}")
        return 0

    if args.command == "extract":
        input_root = Path(args.input).expanduser().resolve()
        output_dir = Path(args.output).expanduser().resolve()
        print(f"Extracting VBA from: {input_root}")
        report = run_extract(input_root, output_dir)
        print(f"Files processed: {report.file_count}")
        print(f"Modules extracted: {report.module_count}")
        print(f"JSON report: {output_dir / 'extract_report.json'}")
        print(f"CSV report: {output_dir / 'extract_report.csv'}")
        return 0

    if args.command == "convert":
        input_dir = Path(args.input).expanduser().resolve()
        output_dir = Path(args.output).expanduser().resolve()
        print(f"Converting VBA → Apps Script from: {input_dir}")
        report = run_convert(input_dir, output_dir, args.api_key or None, args.model)
        print(f"Total: {report.total}  Success: {report.success}  Failed: {report.failed}")
        print(f"JSON report: {output_dir / 'convert_report.json'}")
        print(f"CSV report: {output_dir / 'convert_report.csv'}")
        return 0 if report.failed == 0 else 2

    if args.command == "upload":
        input_root = Path(args.input).expanduser().resolve()
        output_dir = Path(args.output).expanduser().resolve()
        credentials_path = Path(args.credentials).expanduser().resolve()
        token_path = Path(args.token).expanduser().resolve()

        if not credentials_path.exists():
            print(f"Missing credentials file: {credentials_path}")
            return 1

        report = run_upload(
            input_root=input_root,
            output_dir=output_dir,
            credentials_path=credentials_path,
            token_path=token_path,
            drive_folder_id=args.drive_folder_id or None,
            convert_to_sheets=args.convert_to_sheets,
        )
        print(f"Upload attempted: {report.file_count}")
        print(f"Success: {report.success_count}")
        print(f"Failed: {report.failure_count}")
        print(f"JSON report: {output_dir / 'upload_report.json'}")
        print(f"CSV report: {output_dir / 'upload_report.csv'}")
        return 0 if report.failure_count == 0 else 2

    if args.command == "deploy":
        input_dir = Path(args.input).expanduser().resolve()
        credentials_path = Path(args.credentials).expanduser().resolve()
        token_path = Path(args.token).expanduser().resolve()

        if not credentials_path.exists():
            print(f"Missing credentials file: {credentials_path}")
            return 1

        report = run_deploy(
            input_dir=input_dir,
            spreadsheet_id=args.spreadsheet_id,
            credentials_path=credentials_path,
            token_path=token_path,
            script_id=args.script_id or None,
        )
        print(f"Deploy report: {input_dir / 'deploy_report.json'}")
        return 0

    if args.command == "migrate":
        xlsm_path = Path(args.xlsm).expanduser().resolve()
        gs_dir = Path(args.gs_dir).expanduser().resolve()
        output_dir = Path(args.output).expanduser().resolve()
        credentials_path = Path(args.credentials).expanduser().resolve()
        token_path = Path(args.token).expanduser().resolve()

        if not xlsm_path.exists():
            print(f"Missing xlsm file: {xlsm_path}")
            return 1
        if not credentials_path.exists():
            print(f"Missing credentials file: {credentials_path}")
            return 1

        output_dir.mkdir(parents=True, exist_ok=True)

        # Step 1: Upload xlsm as Google Sheets
        print(f"Uploading {xlsm_path.name} to Google Sheets...")
        upload_report = run_upload(
            input_root=xlsm_path,
            output_dir=output_dir,
            credentials_path=credentials_path,
            token_path=token_path,
            drive_folder_id=None,
            convert_to_sheets=True,
        )

        if upload_report.failure_count > 0 or not upload_report.files:
            print("Upload failed.")
            return 1

        uploaded = upload_report.files[0]
        spreadsheet_id = uploaded.drive_file_id
        print(f"Spreadsheet created: {uploaded.drive_web_view_link}")

        # Step 2: Deploy .gs files to the spreadsheet
        print(f"Deploying .gs files from {gs_dir}...")
        deploy_report = run_deploy(
            input_dir=gs_dir,
            spreadsheet_id=spreadsheet_id,
            credentials_path=credentials_path,
            token_path=token_path,
        )
        print(f"Migration complete!")
        print(f"Spreadsheet: {uploaded.drive_web_view_link}")
        print(f"Script editor: https://script.google.com/d/{deploy_report.script_id}/edit")
        return 0

    parser.print_help()
    return 1


if __name__ == "__main__":
    raise SystemExit(main())
